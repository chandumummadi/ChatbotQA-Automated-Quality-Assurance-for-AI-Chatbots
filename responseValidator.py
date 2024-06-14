import sys
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load a pre-trained model
model = SentenceTransformer('paraphrase-MiniLM-L6-v2')

def validate_semantic_similarity(expected, actual, threshold=0.75):
    """
    Validate the semantic similarity between expected and actual answers.
    
    Args:
        expected (str): The expected answer.
        actual (str): The actual answer.
        threshold (float): The similarity threshold for validation.

    Returns:
        bool: True if similarity score is above the threshold, False otherwise.
    """
    embeddings1 = model.encode(expected, convert_to_tensor=True)
    embeddings2 = model.encode(actual, convert_to_tensor=True)
    cosine_scores = util.pytorch_cos_sim(embeddings1, embeddings2)
    print("Cosine Scores:", cosine_scores.item())
    return cosine_scores.item() >= threshold

def process_row(expected_result, actual_result, threshold):
    """
    Process a single row of expected and actual results to validate them.

    Args:
        expected_result (str): The expected result.
        actual_result (str): The actual result.
        threshold (float): The similarity threshold for validation.

    Returns:
        str: "PASS" if the results are similar enough, "FAIL" otherwise.
    """
    if validate_semantic_similarity(expected_result, actual_result, threshold):
        return "PASS"
    else:
        return "FAIL"

def validate_results(file_path, threshold=0.60):
    """
    Validate the results in the Excel file and store validation results.

    Args:
        file_path (str): Path to the Excel file containing expected and actual results.
        threshold (float): The similarity threshold for validation.
    """
    try:
        # Load the Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Ensure the fourth column exists and has a header
        if sheet.max_column < 4:
            sheet.cell(row=1, column=4, value="Validation")

        pass_count = 0
        fail_count = 0

        # Loop through the rows, get expected and actual results, validate and store the results
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=False):
            question_cell, expected_cell, actual_cell = row
            expected_result = expected_cell.value
            actual_result = actual_cell.value

            # Ensure both results are strings
            if isinstance(expected_result, (int, float)):
                expected_result = str(expected_result)
            if isinstance(actual_result, (int, float)):
                actual_result = str(actual_result)

            if expected_result and actual_result:
                validation_status = process_row(expected_result, actual_result, threshold)
                sheet.cell(row=question_cell.row, column=4, value=validation_status)
                print(f"Row {question_cell.row}: Expected: {expected_result}\nActual: {actual_result}\nValidation: {validation_status}\n")
                
                if validation_status == "PASS":
                    pass_count += 1
                else:
                    fail_count += 1
            else:
                print(f"Row {question_cell.row}: Missing expected or actual result")

        # Save the updated workbook
        workbook.save(file_path)
        print(f"Results saved to {file_path}")

        # Plot the pass/fail percentages
        plot_pass_fail_percentage(pass_count, fail_count)

    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def plot_pass_fail_percentage(pass_count, fail_count):
    """
    Plot the percentage of pass/fail test cases.

    Args:
        pass_count (int): The number of pass test cases.
        fail_count (int): The number of fail test cases.
    """
    total = pass_count + fail_count
    pass_percentage = (pass_count / total) * 100
    fail_percentage = (fail_count / total) * 100

    labels = 'Pass', 'Fail'
    sizes = [pass_percentage, fail_percentage]
    colors = ['#4CAF50', '#FF5252']
    explode = (0.1, 0)  # explode the 1st slice (Pass)

    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%',
            shadow=True, startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    plt.title('Pass/Fail Percentage of Test Cases')
    plt.show()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python response_validator.py <file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    validate_results(file_path)
