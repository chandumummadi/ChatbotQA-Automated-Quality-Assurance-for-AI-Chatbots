import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
from openpyxl import load_workbook
from responseValidator import validate_results

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def setup_driver():
    logging.info("Setting up WebDriver")
    return webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

def close_cookie_consent(driver, wait):
    try:
        logging.info("Checking for cookie consent bar")
        cookie_consent = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'cky-consent-bar')))
        close_button = cookie_consent.find_element(By.XPATH, ".//button[contains(text(),'Accept')]")
        close_button.click()
        logging.info("Closed cookie consent bar")
    except Exception as e:
        logging.warning(f"No cookie consent bar found or unable to close it: {e}")

def login(driver, wait):
    try:
        logging.info("Attempting to log in")
        email_input = wait.until(EC.presence_of_element_located((By.NAME, 'email')))
        email_input.send_keys('user@example.com')  # Replace with your actual email

        password_input = wait.until(EC.presence_of_element_located((By.NAME, 'password')))
        password_input.send_keys('Your_password_here')  # Replace with your actual password

        login_button = driver.find_element(By.XPATH, "//button[contains(text(),'Login')]")
        login_button.click()

        wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='textbox' and @aria-multiline='true']")))
        logging.info("Logged in successfully")
    except Exception as e:
        logging.error(f"Error during login: {e}")
        driver.quit()
        exit(1)

def get_chat_input(driver, wait):
    try:
        logging.info("Locating chat input element")
        chat_input = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='textbox' and @aria-multiline='true']")))
        chat_input.click()
        return chat_input
    except Exception as e:
        logging.error(f"Error locating the chat input element: {e}")
        driver.quit()
        exit(1)

def send_message_and_capture_response(driver, chat_input, message):
    chat_input.send_keys("give me most concise and presice answer in a word or a sentence" + message)
    chat_input.send_keys(Keys.RETURN)

    response = None
    initial_message_count = len(driver.find_elements(By.CSS_SELECTOR, "div[data-cy='chatMessage'] div.prose p"))
    while not response:
        time.sleep(1)
        responses = driver.find_elements(By.CSS_SELECTOR, "div[data-cy='chatMessage'] div.prose p")
        if len(responses) > initial_message_count:
            response = responses[-1].text
    return response

def main():
    driver = setup_driver()
    wait = WebDriverWait(driver, 10)

    driver.get('https://app.textcortex.com/user/dashboard/chat')  # replace with the actual URL of Zenochat
    time.sleep(2)  # Ensure the page loads completely

    close_cookie_consent(driver, wait)
    login(driver, wait)

    chat_input = get_chat_input(driver, wait)

    # Load the workbook and select the active worksheet
    workbook = load_workbook('ZenoChatResponses.xlsx')  # Replace with your actual file path
    sheet = workbook.active

    # Ensure there is a third column for responses
    if sheet.max_column < 3:
        sheet.cell(row=1, column=3).value = "Response"

    # Iterate over the rows and read questions from the first column
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=False):
        question = row[0].value
        expected_response = row[1].value
        
        if question:
            response = send_message_and_capture_response(driver, chat_input, question)
            row[2].value = response
            logging.info(f"Asked: {question} | Received: {response}")
            time.sleep(15)

    # Save the workbook with the new responses
    workbook.save('ZenoChatResponses.xlsx')  # Save to the same file

    logging.info("Results saved to questions.xlsx")

    driver.quit()
    logging.info("Script completed and driver closed")
    
    
    validate_respose()
    logging.info("Validation completed")

def validate_respose():
    validate_results('ZenoChatResponses.xlsx')

if __name__ == "__main__":
    main()
    validate_respose()