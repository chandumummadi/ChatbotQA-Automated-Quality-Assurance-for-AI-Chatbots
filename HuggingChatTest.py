import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from openpyxl import load_workbook
from responseValidator import validate_results

# Setup WebDriver options for incognito mode
options = webdriver.ChromeOptions()
options.add_argument("--incognito")

# Setup WebDriver with options
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

# Open Hugging Face Chat Website in incognito mode
driver.get('https://huggingface.co/chat/')

# Wait for the page to load
wait = WebDriverWait(driver, 10)

# Click on "Sign in with Hugging Face" button
try:
    sign_in_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@type='submit' and contains(@class, 'text-lg font-semibold')]")))
    sign_in_button.click()
except Exception as e:
    print(f"Error clicking sign in button: {e}")
    driver.quit()
    exit(1)

# Fill in the login form
try:
    # Locate and fill the username input
    username_input = wait.until(EC.presence_of_element_located((By.NAME, 'username')))
    username_input.send_keys('user@example.com')

    # Locate and fill the password input
    password_input = wait.until(EC.presence_of_element_located((By.NAME, 'password')))
    password_input.send_keys('your_password_here')

    # Locate and click the login button
    login_button = driver.find_element(By.XPATH, "//button[@type='submit' and contains(text(),'Login')]")
    login_button.click()

    # Wait for the chat page to load
    wait.until(EC.presence_of_element_located((By.XPATH, "//textarea[@placeholder='Ask anything']")))
except Exception as e:
    print(f"Error during login: {e}")
    driver.quit()
    exit(1)

# Function to send message and capture response
def send_message_and_capture_response(message):
    try:
        # Re-locate the chat input element each time to avoid stale element reference exception
        chat_input = wait.until(EC.presence_of_element_located((By.XPATH, "//textarea[@placeholder='Ask anything']")))
        chat_input.clear()
        chat_input.send_keys(message)
        chat_input.send_keys(Keys.RETURN)

        # Wait for a new response to appear
        initial_message_count = len(driver.find_elements(By.CSS_SELECTOR, "div.prose"))
        start_time = time.time()
        while time.time() - start_time < 20:  # Timeout of 20 seconds
            responses = driver.find_elements(By.CSS_SELECTOR, "div.prose")
            if len(responses) > initial_message_count:
                break
            print("Waiting for response to stabilise, len of responses", len(responses))
            time.sleep(1)

        # Always pick the last `div.prose` element
        latest_response = driver.find_elements(By.CSS_SELECTOR, "div.prose")[-1]
        response_texts = []
        previous_length = 0

        start_time = time.time()
        # Wait for the response to stabilize with a timeout
        while time.time() - start_time < 15:  # Timeout of 15 seconds
            try:
                # Re-locate the latest `div.prose` element
                latest_response = driver.find_elements(By.CSS_SELECTOR, "div.prose")[-1]
                current_text = "\n".join([child.text for child in latest_response.find_elements(By.XPATH, ".//*") if
                                          child.tag_name in {"p", "br", "td", "th"}])
                if len(current_text) == previous_length:
                    break
                previous_length = len(current_text)
                time.sleep(3)  # Adjust this value to increase the waiting time between checks (in seconds)
            except StaleElementReferenceException:
                continue

        for child in latest_response.find_elements(By.XPATH, ".//*"):
            if child.tag_name == "p":
                response_texts.append(child.text)
            elif child.tag_name == "br":
                response_texts.append("\n")
            print(f"Tag: {child.tag_name}, Text: {child.text}")  # Print tag and text of each child element

        response = "\n".join(response_texts)
        return response
    except Exception as e:
        print(f"Error sending message or capturing response: {e}")
        return "Error capturing response"

# Load the Excel file
workbook = load_workbook('HuggingChatResponses.xlsx')
sheet = workbook.active

# Ensure the third column exists
if sheet.max_column < 3:
    sheet.cell(row=1, column=3, value="Hugging Chat Responses")

# Loop through the rows, get questions, send them to the chat, and store responses
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=False):
    question_cell = row[0]
    question = question_cell.value

    if question:
        response = send_message_and_capture_response("give me most concise and presice answer in a word or a sentence\n" + question)
        sheet.cell(row=question_cell.row, column=3, value=response)  # Write the response to the third column
        print(f"Q: {question}\nA: {response}\n")
        time.sleep(15)

# Save the updated workbook
workbook.save('HuggingChatResponses.xlsx')
print("Results saved to questions.xlsx")

# Close the driver after all questions are sent
driver.quit()


def validate_respose():
    validate_results('HuggingChatResponses.xlsx')


def main():
    validate_respose()
    
if __name__ == "__main__":
    validate_respose()